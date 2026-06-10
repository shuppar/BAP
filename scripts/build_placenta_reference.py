"""
build_placenta_reference.py

Build STAMP reference log2FC matrices for reference-based correlation labelling
of mouse placenta snRNA-seq data at E12.5 and E18.5.

Inputs:
  - mousePlacenta_DEG_AllCells.xlsx  (35 cell types incl. trophoblast + maternal + immune + blood)
  - mousePlacenta_DEG_Tropho.xlsx    (15 trophoblast subtypes, for Phase 7b subcluster pass)

Outputs:
  - refs/stamp_ref_allcells.h5      (35 types x ~6800 genes of avg_log2FC)
  - refs/stamp_ref_tropho.h5        (15 types x ~5000 genes of avg_log2FC)
  - refs/stamp_ref_summary.csv      (per-type n_DEGs + diagnostics)

DEG filter: padj < 0.05 & |log2FC| > 0.5
Non-DEG genes filled with 0 (vector is "expression deviation vs other types").
Rank-transformed at annotation time (Spearman), not here.

Run on workstation (or local Mac — input is small).
"""
from __future__ import annotations
import argparse
from pathlib import Path
import sys

import pandas as pd
import numpy as np
from openpyxl import load_workbook


# ---------------------------- helpers --------------------------------------- #

def load_deg_xlsx(path: Path) -> pd.DataFrame:
    """Load all sheets of a STAMP DEG xlsx into a single long DataFrame."""
    wb = load_workbook(path, read_only=True, data_only=True)
    frames = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        df = pd.DataFrame(rows[1:], columns=rows[0])
        df = df[df["gene"].notna()].copy()
        # `cluster` column is supposed to be there; trust it but verify
        if df["cluster"].isna().all():
            df["cluster"] = sheet
        frames.append(df)
    wb.close()
    out = pd.concat(frames, ignore_index=True)
    for c in ["p_val", "avg_log2FC", "pct.1", "pct.2", "p_val_adj"]:
        out[c] = pd.to_numeric(out[c], errors="coerce")
    return out


def build_reference_matrix(
    deg: pd.DataFrame,
    padj_max: float = 0.05,
    log2fc_min: float = 0.5,
    pct1_min: float = 0.10,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Return (ref_matrix, diagnostics).

    ref_matrix: ref_types (rows) x genes (cols), values = avg_log2FC for that
    (type, gene) pair. Non-DEG (type, gene) entries are filled with 0.

    diagnostics: per-type DEG counts.
    """
    # Filter to significantly-up DEGs only — these are what discriminate types.
    # Downregulated genes carry information too, but they're typically just
    # the inverse signal of another type's up-DEGs; including them would
    # double-count and they're often less reliable in droplet snRNA-seq.
    sig = deg[
        (deg["p_val_adj"] < padj_max)
        & (deg["avg_log2FC"] > log2fc_min)
        & (deg["pct.1"] >= pct1_min)
    ].copy()

    # Per (type, gene), if there are duplicates (shouldn't be, but defensive),
    # keep the highest log2FC.
    sig = (
        sig.sort_values("avg_log2FC", ascending=False)
        .drop_duplicates(subset=["cluster", "gene"], keep="first")
    )

    # Pivot: rows = cell type, cols = gene, values = log2FC
    ref_mat = (
        sig.pivot(index="cluster", columns="gene", values="avg_log2FC")
        .fillna(0.0)
    )

    # Diagnostics — per-type n_DEGs and median FC
    diag = (
        sig.groupby("cluster")
        .agg(
            n_DEGs=("gene", "nunique"),
            median_log2FC=("avg_log2FC", "median"),
            median_pct1=("pct.1", "median"),
            median_pct2=("pct.2", "median"),
        )
        .reset_index()
        .rename(columns={"cluster": "cell_type"})
    )

    # How specific is each type's signature? Count genes with very high FC
    # (>2) and low pct.2 (<0.20) — these are the "near-unique" markers.
    near_unique = (
        sig[(sig["avg_log2FC"] > 2.0) & (sig["pct.2"] < 0.20)]
        .groupby("cluster")["gene"]
        .nunique()
        .rename("n_near_unique_DEGs")
    )
    diag = diag.merge(
        near_unique.reset_index().rename(columns={"cluster": "cell_type"}),
        on="cell_type",
        how="left",
    )
    diag["n_near_unique_DEGs"] = diag["n_near_unique_DEGs"].fillna(0).astype(int)

    return ref_mat, diag


def write_h5(mat: pd.DataFrame, out: Path) -> None:
    """Write reference matrix to HDF5. Loadable via load_reference() below."""
    import h5py
    out.parent.mkdir(parents=True, exist_ok=True)
    with h5py.File(out, "w") as f:
        f.create_dataset("log2fc", data=mat.values.astype(np.float32), compression="gzip")
        f.create_dataset(
            "cell_types",
            data=np.array(mat.index.astype(str), dtype=h5py.string_dtype()),
        )
        f.create_dataset(
            "genes",
            data=np.array(mat.columns.astype(str), dtype=h5py.string_dtype()),
        )


def load_reference(path: Path) -> pd.DataFrame:
    """Companion loader — used by the annotation script."""
    import h5py
    with h5py.File(path, "r") as f:
        mat = f["log2fc"][:]
        types = [s.decode() if isinstance(s, bytes) else s for s in f["cell_types"][:]]
        genes = [s.decode() if isinstance(s, bytes) else s for s in f["genes"][:]]
    return pd.DataFrame(mat, index=pd.Index(types, name="cell_type"), columns=genes)


# ---------------------------- main ------------------------------------------ #

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--allcells-xlsx",
        type=Path,
        required=True,
        help="Path to mousePlacenta_DEG_AllCells.xlsx",
    )
    ap.add_argument(
        "--tropho-xlsx",
        type=Path,
        required=True,
        help="Path to mousePlacenta_DEG_Tropho.xlsx",
    )
    ap.add_argument(
        "--out-dir",
        type=Path,
        default=Path("refs"),
        help="Output directory (default: refs/)",
    )
    ap.add_argument("--padj-max", type=float, default=0.05)
    ap.add_argument("--log2fc-min", type=float, default=0.5)
    ap.add_argument("--pct1-min", type=float, default=0.10)
    args = ap.parse_args()

    for p in (args.allcells_xlsx, args.tropho_xlsx):
        if not p.exists():
            print(f"ERROR: {p} not found", file=sys.stderr)
            return 1

    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    diagnostics_all = []

    for name, xlsx in [
        ("allcells", args.allcells_xlsx),
        ("tropho", args.tropho_xlsx),
    ]:
        print(f"\n=== {name} ===")
        print(f"Loading {xlsx} ...")
        deg = load_deg_xlsx(xlsx)
        print(f"  loaded: {len(deg):,} DEG rows across {deg['cluster'].nunique()} cell types")

        print(
            f"Filtering: padj<{args.padj_max}  |log2FC|>{args.log2fc_min}  "
            f"pct.1>={args.pct1_min}"
        )
        mat, diag = build_reference_matrix(
            deg,
            padj_max=args.padj_max,
            log2fc_min=args.log2fc_min,
            pct1_min=args.pct1_min,
        )
        print(f"  reference matrix: {mat.shape[0]} cell types x {mat.shape[1]:,} genes")
        print(f"  sparsity: {(mat == 0).mean().mean() * 100:.1f}% zeros")

        diag["source"] = name
        diagnostics_all.append(diag)

        out_h5 = out_dir / f"stamp_ref_{name}.h5"
        write_h5(mat, out_h5)
        print(f"  wrote {out_h5}")

    diag_df = pd.concat(diagnostics_all, ignore_index=True)
    diag_path = out_dir / "stamp_ref_summary.csv"
    diag_df.to_csv(diag_path, index=False)
    print(f"\nDiagnostics written to {diag_path}")
    print("\nPer-type DEG counts:")
    print(diag_df.to_string(index=False))

    # Quick sanity warnings
    weak = diag_df[diag_df["n_near_unique_DEGs"] < 5]
    if not weak.empty:
        print(
            "\nWARN: cell types with <5 near-unique markers "
            "(may be hard to discriminate):"
        )
        print(weak[["cell_type", "source", "n_DEGs", "n_near_unique_DEGs"]].to_string(index=False))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
